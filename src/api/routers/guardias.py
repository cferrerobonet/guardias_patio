"""
API REST Router - Guardias

Endpoints para gestión de guardias (consultar, generar, asignar).
"""

import csv
import io
from datetime import date
from time import perf_counter
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from api.dependencies import get_db
from application.dtos import FiltroGuardiasDTO
from application.dtos.guardia_dto import CrearGuardiaDTO
from application.use_cases.guardia.asignar_guardia import AsignarGuardiaUseCase
from application.use_cases.guardia.limpiar_guardias import LimpiarGuardiasUseCase
from application.use_cases.guardia.obtener_guardias import ObtenerGuardiasUseCase
from core.exceptions import BusinessLogicError, ValidationError
from core.logging import get_logger
from infrastructure.repositories import SQLAlchemyGuardiaRepository

router = APIRouter(prefix="/guardias", tags=["guardias"])
logger = get_logger(__name__)


class GuardiaResponse(BaseModel):
    """Schema de respuesta para guardia."""

    id: int
    fecha: date
    recreo: int
    turno: str
    zona_id: int
    zona_nombre: Optional[str] = None
    profesor_id: Optional[int] = None
    profesor_nombre: Optional[str] = None
    es_sustitucion: bool = False

    class Config:
        from_attributes = True


class GuardiasCountResponse(BaseModel):
    total: int


class PaginatedGuardiasResponse(BaseModel):
    """Respuesta paginada de guardias."""

    items: List[GuardiaResponse]
    total: int
    page: int
    size: int
    pages: int


@router.get("", response_model=PaginatedGuardiasResponse, summary="Listar guardias")
def obtener_guardias(
    configuracion_id: int,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    profesor_id: Optional[int] = None,
    zona_id: Optional[int] = None,
    turno: Optional[str] = None,
    limit: int = Query(default=100, le=1000),
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """
    Obtiene guardias con filtros opcionales.

    Args:
        configuracion_id: ID de la configuración del curso
        fecha_inicio: Filtrar desde fecha (opcional)
        fecha_fin: Filtrar hasta fecha (opcional)
        profesor_id: Filtrar por profesor (opcional)
        zona_id: Filtrar por zona (opcional)
        turno: Filtrar por turno (opcional)
        limit: Máximo de resultados (default: 100, max: 1000)
        offset: Desplazamiento para paginación
        db: Sesión de base de datos (inyectada)

    Returns:
        List[GuardiaResponse]: Lista de guardias

    Examples:
        GET /api/guardias?configuracion_id=1&turno=mañana&limit=50
    """
    start = perf_counter()
    try:
        filtros = FiltroGuardiasDTO(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            profesor_id=profesor_id,
            zona_id=zona_id,
            turno=turno,
        )
        use_case = ObtenerGuardiasUseCase(db)
        dtos = use_case.execute(filtros)

        # Filtrar por curso y paginar
        dtos = [g for g in dtos if True]  # curso_id se filtra en el use case si se añade
        total = len(dtos)
        paginados = dtos[offset: offset + limit]
        page = (offset // limit) + 1 if limit > 0 else 1
        pages = (total + limit - 1) // limit if limit > 0 else 1

        items = [
            GuardiaResponse(
                id=g.id,
                fecha=g.fecha,
                recreo=g.numero_recreo,
                turno=g.turno,
                zona_id=g.zona_id,
                zona_nombre=g.zona_nombre,
                profesor_id=g.profesor_id,
                profesor_nombre=g.profesor_nombre,
                es_sustitucion=g.es_sustitucion,
            )
            for g in paginados
        ]

        duration_ms = (perf_counter() - start) * 1000
        logger.info(
            "Metrica negocio: listado guardias",
            extra={
                "metric": "guardias_listado",
                "total": total,
                "devueltas": len(items),
                "page": page,
                "offset": offset,
                "limit": limit,
                "duration_ms": round(duration_ms, 2),
            },
        )
        return PaginatedGuardiasResponse(items=items, total=total, page=page, size=len(items), pages=pages)

    except (SQLAlchemyError, ValueError, TypeError, OSError):
        logger.exception("Error en listado de guardias")
        raise HTTPException(status_code=500, detail="Error al obtener guardias")


@router.get("/count", response_model=GuardiasCountResponse, summary="Contar guardias")
def contar_guardias(
    configuracion_id: int,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    profesor_id: Optional[int] = None,
    zona_id: Optional[int] = None,
    turno: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Cuenta guardias con filtros opcionales.

    Args:
        configuracion_id: ID de la configuración del curso
        fecha_inicio: Filtrar desde fecha (opcional)
        fecha_fin: Filtrar hasta fecha (opcional)
        profesor_id: Filtrar por profesor (opcional)
        zona_id: Filtrar por zona (opcional)
        turno: Filtrar por turno (opcional)
        db: Sesión de base de datos (inyectada)

    Returns:
        dict: Total de guardias

    Examples:
        GET /api/guardias/count?configuracion_id=1&turno=tarde
    """
    start = perf_counter()
    try:
        filtros = FiltroGuardiasDTO(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            profesor_id=profesor_id,
            zona_id=zona_id,
            turno=turno,
        )
        use_case = ObtenerGuardiasUseCase(db)
        dtos = use_case.execute(filtros)
        total = len(dtos)
        duration_ms = (perf_counter() - start) * 1000
        logger.info(
            "Metrica negocio: conteo guardias",
            extra={
                "metric": "guardias_conteo",
                "total": total,
                "duration_ms": round(duration_ms, 2),
            },
        )
        return {"total": total}

    except (SQLAlchemyError, ValueError, TypeError, OSError):
        logger.exception("Error en conteo de guardias")
        raise HTTPException(status_code=500, detail="Error al contar guardias")


def _get_guardias_dtos(
    configuracion_id: int,
    fecha_inicio: Optional[date],
    fecha_fin: Optional[date],
    profesor_id: Optional[int],
    zona_id: Optional[int],
    turno: Optional[str],
    db: Session,
):
    filtros = FiltroGuardiasDTO(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        profesor_id=profesor_id,
        zona_id=zona_id,
        turno=turno,
    )
    use_case = ObtenerGuardiasUseCase(db)
    return use_case.execute(filtros)


_EXPORT_COLUMNS = ["id", "fecha", "recreo", "turno", "zona_id", "zona_nombre", "profesor_id", "profesor_nombre", "es_sustitucion"]


def _dto_to_row(g) -> list:
    return [g.id, g.fecha, g.numero_recreo, g.turno, g.zona_id, g.zona_nombre, g.profesor_id, g.profesor_nombre, g.es_sustitucion]


@router.get("/export/csv", response_class=Response, summary="Exportar guardias a CSV")
def exportar_guardias_csv(
    configuracion_id: int,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    profesor_id: Optional[int] = None,
    zona_id: Optional[int] = None,
    turno: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Exporta las guardias filtradas como archivo CSV (UTF-8 con BOM para Excel)."""
    start = perf_counter()
    try:
        dtos = _get_guardias_dtos(configuracion_id, fecha_inicio, fecha_fin, profesor_id, zona_id, turno, db)
    except (ValueError, TypeError, OSError):
        logger.exception("Error al exportar guardias CSV")
        raise HTTPException(status_code=500, detail="Error al obtener guardias")

    buf = io.StringIO()
    buf.write("\ufeff")  # BOM UTF-8 para compatibilidad con Excel
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(_EXPORT_COLUMNS)
    for g in dtos:
        writer.writerow(_dto_to_row(g))

    content = buf.getvalue().encode("utf-8")
    duration_ms = (perf_counter() - start) * 1000
    logger.info(
        "Metrica negocio: exportacion guardias CSV",
        extra={
            "metric": "guardias_export_csv",
            "registros": len(dtos),
            "duration_ms": round(duration_ms, 2),
        },
    )
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=guardias.csv"},
    )


@router.get("/export/xlsx", summary="Exportar guardias a Excel")
def exportar_guardias_xlsx(
    configuracion_id: int,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    profesor_id: Optional[int] = None,
    zona_id: Optional[int] = None,
    turno: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Exporta las guardias filtradas como archivo Excel (.xlsx)."""
    start = perf_counter()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no disponible")

    try:
        dtos = _get_guardias_dtos(configuracion_id, fecha_inicio, fecha_fin, profesor_id, zona_id, turno, db)
    except (ValueError, TypeError, OSError):
        logger.exception("Error al exportar guardias XLSX")
        raise HTTPException(status_code=500, detail="Error al obtener guardias")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guardias"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

    ws.append(_EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for g in dtos:
        ws.append(_dto_to_row(g))

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    duration_ms = (perf_counter() - start) * 1000
    logger.info(
        "Metrica negocio: exportacion guardias XLSX",
        extra={
            "metric": "guardias_export_xlsx",
            "registros": len(dtos),
            "duration_ms": round(duration_ms, 2),
        },
    )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=guardias.xlsx"},
    )


@router.post("", response_model=GuardiaResponse, status_code=201, summary="Asignar guardia")
def asignar_guardia(guardia: CrearGuardiaDTO, db: Session = Depends(get_db)):
    """Asigna manualmente una guardia a un profesor en una zona y fecha concretas."""
    start = perf_counter()
    try:
        dto = AsignarGuardiaUseCase(db).execute(guardia)
        duration_ms = (perf_counter() - start) * 1000
        logger.info(
            "Metrica negocio: guardia asignada",
            extra={
                "metric": "guardia_asignada",
                "guardia_id": dto.id,
                "profesor_id": dto.profesor_id,
                "zona_id": dto.zona_id,
                "turno": dto.turno,
                "duration_ms": round(duration_ms, 2),
            },
        )
        return GuardiaResponse(
            id=dto.id,
            fecha=dto.fecha,
            recreo=dto.numero_recreo,
            turno=dto.turno,
            zona_id=dto.zona_id,
            zona_nombre=dto.zona_nombre,
            profesor_id=dto.profesor_id,
            profesor_nombre=dto.profesor_nombre,
            es_sustitucion=dto.es_sustitucion,
        )
    except (ValidationError, BusinessLogicError) as e:
        raise HTTPException(status_code=422, detail={"code": "validation_error", "message": str(e)})
    except (SQLAlchemyError, ValueError, TypeError, OSError) as e:
        logger.exception("Error al asignar guardia")
        raise HTTPException(status_code=500, detail={"code": "internal_error", "message": str(e)})


@router.delete("", status_code=200, summary="Eliminar todas las guardias")
def limpiar_guardias(db: Session = Depends(get_db)):
    """Elimina TODAS las guardias del sistema. Operación irreversible."""
    start = perf_counter()
    try:
        repo = SQLAlchemyGuardiaRepository(db)
        total = LimpiarGuardiasUseCase(repo).execute()
        duration_ms = (perf_counter() - start) * 1000
        logger.info(
            "Metrica negocio: limpieza de guardias",
            extra={
                "metric": "guardias_limpieza",
                "eliminadas": total,
                "duration_ms": round(duration_ms, 2),
            },
        )
        return {"eliminadas": total}
    except (ValueError, OSError) as e:
        logger.exception("Error al limpiar guardias")
        raise HTTPException(status_code=500, detail={"code": "internal_error", "message": str(e)})

